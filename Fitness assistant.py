"""
Offline AI Workout Plan Generator — with Chat
==============================================
Generates your personalised workout plan, displays
it as tables, saves to Excel, then opens a chat
where you can ask fitness questions with full context.

Requirements:
    pip install ollama rich openpyxl

Setup:
    1. Install Ollama from https://ollama.com
    2. ollama pull llama3.2:3b
    3. pip install ollama rich openpyxl
    4. python workout_recommender_chat.py
"""

import ollama
import json
import re
from datetime import datetime
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.text import Text
from rich import box
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

console = Console()

# ─────────────────────────────────────────────
# STEP 1: Collect user input
# ─────────────────────────────────────────────

def get_user_profile():
    console.print("\n" + "=" * 54)
    console.print("   [bold]  OFFLINE AI WORKOUT PLAN GENERATOR[/bold]")
    console.print("        Powered by Llama 3.2 (local)")
    console.print("=" * 54)
    console.print("Answer a few questions to get your personalised plan.\n")

    name = input("Your name: ").strip() or "Athlete"

    while True:
        try:
            age = int(input("Age (years): "))
            if 10 <= age <= 80: break
            print("  ⚠  Enter age between 10 and 80.")
        except ValueError:
            print("  ⚠  Please enter a valid number.")

    while True:
        try:
            height_cm = float(input("Height (cm): "))
            if 100 <= height_cm <= 250: break
            print("  ⚠  Enter a realistic height in cm.")
        except ValueError:
            print("  ⚠  Please enter a valid number.")

    while True:
        try:
            weight_kg = float(input("Weight (kg): "))
            if 30 <= weight_kg <= 250: break
            print("  ⚠  Enter a realistic weight in kg.")
        except ValueError:
            print("  ⚠  Please enter a valid number.")

    print("\nExperience level:")
    print("  1. Beginner      (0–6 months)")
    print("  2. Intermediate  (6 months – 2 years)")
    print("  3. Advanced      (2+ years)")
    while True:
        c = input("Choose (1/2/3): ").strip()
        exp_map = {"1": "Beginner", "2": "Intermediate", "3": "Advanced"}
        if c in exp_map: experience = exp_map[c]; break
        print("  ⚠  Enter 1, 2, or 3.")

    print("\nExercise style:")
    print("  1. Calisthenics only")
    print("  2. Weighted only (gym)")
    print("  3. Mixed (both)")
    while True:
        c = input("Choose (1/2/3): ").strip()
        style_map = {
            "1": "Calisthenics (bodyweight only)",
            "2": "Weighted gym training",
            "3": "Mixed (calisthenics + weighted)"
        }
        if c in style_map: style = style_map[c]; break
        print("  ⚠  Enter 1, 2, or 3.")

    print("\nPrimary goal:")
    print("  1. Build muscle")
    print("  2. Lose fat")
    print("  3. Build strength")
    print("  4. Improve endurance")
    print("  5. General fitness")
    while True:
        c = input("Choose (1–5): ").strip()
        goal_map = {
            "1": "Build muscle (hypertrophy)",
            "2": "Lose fat and get lean",
            "3": "Build raw strength",
            "4": "Improve endurance and stamina",
            "5": "General fitness and stay active"
        }
        if c in goal_map: goal = goal_map[c]; break
        print("  ⚠  Enter 1 to 5.")

    while True:
        try:
            days = int(input("\nTraining days per week (1–7): "))
            if 1 <= days <= 7: break
            print("  ⚠  Enter between 1 and 7.")
        except ValueError:
            print("  ⚠  Please enter a valid number.")

    print("\nEquipment available:")
    print("  1. None (home / park only)")
    print("  2. Pull-up bar + resistance bands")
    print("  3. Full gym access")
    while True:
        c = input("Choose (1/2/3): ").strip()
        equip_map = {
            "1": "No equipment (bodyweight only)",
            "2": "Pull-up bar and resistance bands",
            "3": "Full gym (barbells, dumbbells, machines)"
        }
        if c in equip_map: equipment = equip_map[c]; break
        print("  ⚠  Enter 1, 2, or 3.")

    return {
        "name": name, "age": age,
        "height_cm": height_cm, "weight_kg": weight_kg,
        "experience": experience, "style": style,
        "goal": goal, "days_per_week": days,
        "equipment": equipment
    }


# ─────────────────────────────────────────────
# STEP 2: BMI analysis
# ─────────────────────────────────────────────

def analyse_profile(profile):
    bmi = round(profile["weight_kg"] / ((profile["height_cm"] / 100) ** 2), 1)
    if bmi < 18.5:   cat, note = "Underweight", "Prioritise caloric surplus and compound lifts."
    elif bmi < 25:   cat, note = "Normal weight", "Great baseline. Tailor tightly to your goal."
    elif bmi < 30:   cat, note = "Overweight",  "Include cardio. Monitor diet alongside training."
    else:            cat, note = "Obese",        "Start low-impact. Build consistency before intensity."
    return {"bmi": bmi, "category": cat, "note": note}


# ─────────────────────────────────────────────
# STEP 3: Build plan-generation prompt (JSON)
# ─────────────────────────────────────────────

def build_plan_prompt(profile, analysis):
    return f"""You are an expert fitness coach. Create a personalised weekly workout plan.

User profile:
- Name: {profile['name']}, Age: {profile['age']}
- Height: {profile['height_cm']}cm, Weight: {profile['weight_kg']}kg, BMI: {analysis['bmi']} ({analysis['category']})
- Experience: {profile['experience']}
- Style: {profile['style']}
- Goal: {profile['goal']}
- Training days/week: {profile['days_per_week']}
- Equipment: {profile['equipment']}
- Coach note: {analysis['note']}

Return ONLY a valid JSON object. No explanation. No markdown. No code fences.
Use exactly this structure:

{{
  "weekly_plan": [
    {{
      "day": "Day 1",
      "focus": "Chest & Triceps",
      "type": "Training",
      "exercises": [
        {{"exercise": "Push-ups", "sets": "4", "reps_duration": "12 reps", "rest": "60 sec"}}
      ]
    }},
    {{
      "day": "Day 2",
      "focus": "Rest",
      "type": "Rest",
      "exercises": [
        {{"exercise": "Light walk or stretching", "sets": "-", "reps_duration": "20 min", "rest": "-"}}
      ]
    }}
  ],
  "warmup": [
    {{"exercise": "Jumping jacks", "sets": "1", "reps_duration": "2 min", "rest": "-"}}
  ],
  "cooldown": [
    {{"exercise": "Standing quad stretch", "sets": "1", "reps_duration": "30 sec each", "rest": "-"}}
  ],
  "tips": ["Tip 1", "Tip 2", "Tip 3"]
}}

Fill all {profile['days_per_week']} training days plus rest days. Be specific with exercise names.
""".strip()


# ─────────────────────────────────────────────
# STEP 4: Generate plan via Ollama
# ─────────────────────────────────────────────

def generate_plan(prompt):
    console.print("\n[bold]  Generating your plan (running locally)...[/bold]")
    console.print("  This may take 20–40 seconds...\n")

    try:
        response = ollama.chat(
            model="llama3.2:3b",
            messages=[{"role": "user", "content": prompt}],
            stream=False
        )
        raw = response["message"]["content"].strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        match = re.search(r"\{.*\}", raw, re.DOTALL)
        if match:
            raw = match.group()
        return json.loads(raw)

    except json.JSONDecodeError:
        console.print("[red]❌  Could not parse JSON. Try running again.[/red]")
        return None
    except Exception as e:
        err = str(e).lower()
        console.print("\n[red]❌  Error connecting to Ollama.[/red]")
        if "connection" in err:
            console.print("   → Ollama is not running. Open the Ollama app first.")
        elif "model" in err:
            console.print("   → Run: ollama pull llama3.2:3b")
        else:
            console.print(f"   → {e}")
        return None


# ─────────────────────────────────────────────
# STEP 5: Display tables in terminal
# ─────────────────────────────────────────────

def display_terminal_tables(profile, analysis, data):
    # Profile summary
    console.print("\n")
    pt = Table(title="  Your Profile", box=box.ROUNDED,
               title_style="bold", header_style="bold white on dark_blue")
    pt.add_column("Field",  style="cyan", width=18)
    pt.add_column("Value",  style="white")
    pt.add_row("Name",           profile["name"])
    pt.add_row("Age",            f"{profile['age']} years")
    pt.add_row("Height / Weight",f"{profile['height_cm']} cm  /  {profile['weight_kg']} kg")
    pt.add_row("BMI",            f"{analysis['bmi']}  ({analysis['category']})")
    pt.add_row("Experience",     profile["experience"])
    pt.add_row("Style",          profile["style"])
    pt.add_row("Goal",           profile["goal"])
    pt.add_row("Training days",  f"{profile['days_per_week']} days/week")
    pt.add_row("Equipment",      profile["equipment"])
    console.print(pt)

    # Warm-up
    console.print("\n")
    wt = Table(title="🔥  Warm-Up Routine", box=box.SIMPLE_HEAVY,
               title_style="bold yellow", header_style="bold yellow")
    wt.add_column("Exercise",        style="white", min_width=25)
    wt.add_column("Sets",            justify="center", width=6)
    wt.add_column("Duration / Reps", justify="center", width=18)
    wt.add_column("Rest",            justify="center", width=10)
    for ex in data.get("warmup", []):
        wt.add_row(ex.get("exercise",""), ex.get("sets","-"),
                   ex.get("reps_duration","-"), ex.get("rest","-"))
    console.print(wt)

    # Daily plan
    for day_data in data.get("weekly_plan", []):
        is_rest = day_data.get("type", "Training") == "Rest"
        color   = "dim" if is_rest else "green"
        console.print(f"\n[bold {color}]{day_data['day']}  —  {day_data['focus']}[/bold {color}]")
        dt = Table(box=box.SIMPLE, show_header=True, header_style="bold")
        dt.add_column("Exercise",        style="white", min_width=28)
        dt.add_column("Sets",            justify="center", width=6)
        dt.add_column("Reps / Duration", justify="center", width=18)
        dt.add_column("Rest",            justify="center", width=12)
        for ex in day_data.get("exercises", []):
            dt.add_row(ex.get("exercise",""), ex.get("sets","-"),
                       ex.get("reps_duration","-"), ex.get("rest","-"))
        console.print(dt)

    # Cool-down
    console.print("\n")
    ct = Table(title="❄️   Cool-Down Routine", box=box.SIMPLE_HEAVY,
               title_style="bold cyan", header_style="bold cyan")
    ct.add_column("Exercise",        style="white", min_width=25)
    ct.add_column("Sets",            justify="center", width=6)
    ct.add_column("Duration / Reps", justify="center", width=18)
    ct.add_column("Rest",            justify="center", width=10)
    for ex in data.get("cooldown", []):
        ct.add_row(ex.get("exercise",""), ex.get("sets","-"),
                   ex.get("reps_duration","-"), ex.get("rest","-"))
    console.print(ct)

    # Tips
    console.print("\n[bold]💡  Personalised Tips[/bold]")
    for i, tip in enumerate(data.get("tips", []), 1):
        console.print(f"  {i}. {tip}")


# ─────────────────────────────────────────────
# STEP 6: Save to Excel
# ─────────────────────────────────────────────

def save_excel(profile, analysis, data):
    wb    = Workbook()
    DARK  = "1F3864"; MID = "2F5496"; LBLUE = "D6E4F0"
    YELL  = "FFF2CC"; GREN = "E2EFDA"; WHITE = "FFFFFF"; GREY = "F2F2F2"
    thin  = Side(style="thin", color="AAAAAA")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hf(sz=11):   return Font(name="Arial", size=sz, bold=True,  color="FFFFFF")
    def cf(bold=False): return Font(name="Arial", size=10, bold=bold, color="000000")
    def fl(c):       return PatternFill("solid", fgColor=c)
    def ctr():       return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lft():       return Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def sc(cell, font=None, bg=None, align=None):
        if font:  cell.font      = font
        if bg:    cell.fill      = fl(bg)
        if align: cell.alignment = align
        cell.border = bdr

    # Sheet 1 — Profile
    ws1 = wb.active; ws1.title = "Profile"
    ws1.merge_cells("A1:C1")
    ws1["A1"] = f"Workout Plan — {profile['name']}"
    ws1["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws1["A1"].fill = fl(DARK); ws1["A1"].alignment = ctr()
    ws1.row_dimensions[1].height = 32
    ws1.merge_cells("A2:C2")
    ws1["A2"] = f"Generated on {datetime.now().strftime('%d %B %Y')}"
    ws1["A2"].font = Font(name="Arial", size=9, color="FFFFFF")
    ws1["A2"].fill = fl(MID); ws1["A2"].alignment = ctr()
    for col, h in enumerate(["Field","Value","Notes"], 1):
        c = ws1.cell(row=3, column=col, value=h); sc(c, hf(), DARK, ctr())
    rows = [
        ("Name",          profile["name"],              ""),
        ("Age",           f"{profile['age']} years",    ""),
        ("Height",        f"{profile['height_cm']} cm", ""),
        ("Weight",        f"{profile['weight_kg']} kg", ""),
        ("BMI",           str(analysis["bmi"]),          analysis["category"]),
        ("Experience",    profile["experience"],          ""),
        ("Style",         profile["style"],               ""),
        ("Goal",          profile["goal"],                ""),
        ("Training Days", f"{profile['days_per_week']} days/week", ""),
        ("Equipment",     profile["equipment"],           ""),
        ("Coach Note",    analysis["note"],               ""),
    ]
    for r, (f, v, n) in enumerate(rows, 4):
        bg = GREY if r % 2 == 0 else WHITE
        for col, txt in enumerate([f, v, n], 1):
            c = ws1.cell(row=r, column=col, value=txt)
            sc(c, cf(bold=(col==1)), bg, lft())
    for col, w in zip(["A","B","C"], [18, 32, 38]):
        ws1.column_dimensions[col].width = w

    # Sheet 2 — Warm-up & Cool-down
    ws2 = wb.create_sheet("Warm-up & Cool-down")
    def write_section(ws, start, title, exercises, tc, rc):
        ws.merge_cells(f"A{start}:D{start}")
        ws[f"A{start}"] = title
        ws[f"A{start}"].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        ws[f"A{start}"].fill = fl(tc); ws[f"A{start}"].alignment = ctr()
        ws.row_dimensions[start].height = 24
        for col, h in enumerate(["Exercise","Sets","Duration / Reps","Rest"], 1):
            c = ws.cell(row=start+1, column=col, value=h); sc(c, hf(10), MID, ctr())
        for i, ex in enumerate(exercises):
            row = start + 2 + i
            bg = rc if i % 2 == 0 else WHITE
            for col, v in enumerate([ex.get("exercise",""), ex.get("sets","-"),
                                      ex.get("reps_duration","-"), ex.get("rest","-")], 1):
                c = ws.cell(row=row, column=col, value=v)
                sc(c, cf(), bg, lft() if col==1 else ctr())
        return start + 2 + len(exercises) + 2
    nr = write_section(ws2, 1,   "🔥  Warm-Up Routine",   data.get("warmup",[]),   "B7722A", YELL)
    write_section(     ws2, nr,  "❄️  Cool-Down Routine", data.get("cooldown",[]), "1F5F8B", LBLUE)
    for col, w in zip(["A","B","C","D"], [30, 8, 20, 12]):
        ws2.column_dimensions[col].width = w

    # Sheet 3 — Weekly Plan
    ws3 = wb.create_sheet("Weekly Plan")
    ws3.merge_cells("A1:E1")
    ws3["A1"] = "Weekly Workout Schedule"
    ws3["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    ws3["A1"].fill = fl(DARK); ws3["A1"].alignment = ctr()
    ws3.row_dimensions[1].height = 28
    cr = 2
    for day_data in data.get("weekly_plan", []):
        is_rest   = day_data.get("type","Training") == "Rest"
        day_color = "888888" if is_rest else "4A7C59"
        ws3.merge_cells(f"A{cr}:E{cr}")
        ws3[f"A{cr}"] = f"{day_data['day']}  —  {day_data['focus']}"
        ws3[f"A{cr}"].font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        ws3[f"A{cr}"].fill = fl(day_color); ws3[f"A{cr}"].alignment = lft()
        ws3.row_dimensions[cr].height = 22; cr += 1
        for col, h in enumerate(["Exercise","Sets","Reps / Duration","Rest","Notes"], 1):
            c = ws3.cell(row=cr, column=col, value=h); sc(c, hf(9), MID, ctr())
        cr += 1
        for i, ex in enumerate(day_data.get("exercises",[])):
            bg = GREN if (not is_rest and i%2==0) else (GREY if is_rest else WHITE)
            for col, v in enumerate([ex.get("exercise",""), ex.get("sets","-"),
                                      ex.get("reps_duration","-"), ex.get("rest","-"), ""], 1):
                c = ws3.cell(row=cr, column=col, value=v)
                sc(c, cf(), bg, lft() if col==1 else ctr())
            cr += 1
        cr += 1
    for col, w in zip(["A","B","C","D","E"], [30, 8, 20, 12, 22]):
        ws3.column_dimensions[col].width = w

    # Sheet 4 — Tips
    ws4 = wb.create_sheet("Tips")
    ws4.merge_cells("A1:B1")
    ws4["A1"] = "  Personalised Coaching Tips"
    ws4["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    ws4["A1"].fill = fl(DARK); ws4["A1"].alignment = ctr()
    ws4.row_dimensions[1].height = 28
    for i, tip in enumerate(data.get("tips",[]), 1):
        bg = LBLUE if i%2==0 else WHITE
        c1 = ws4.cell(row=i+1, column=1, value=f"Tip {i}")
        c2 = ws4.cell(row=i+1, column=2, value=tip)
        sc(c1, cf(bold=True), bg, ctr())
        sc(c2, cf(),          bg, lft())
        ws4.row_dimensions[i+1].height = 36
    ws4.column_dimensions["A"].width = 10
    ws4.column_dimensions["B"].width = 70

    filename = f"{profile['name'].replace(' ','_')}_workout_plan.xlsx"
    wb.save(filename)
    return filename


# ─────────────────────────────────────────────
# STEP 7: Build the system prompt for chat
#         This gives the model full context so
#         every answer is personalised to YOU
# ─────────────────────────────────────────────

def build_system_prompt(profile, analysis, data):
    # Flatten the weekly plan into readable text for context
    plan_text = ""
    for day in data.get("weekly_plan", []):
        plan_text += f"\n{day['day']} — {day['focus']} ({day['type']})\n"
        for ex in day.get("exercises", []):
            plan_text += f"  • {ex['exercise']}  {ex['sets']} sets x {ex['reps_duration']}  rest: {ex['rest']}\n"

    tips_text = "\n".join(f"- {t}" for t in data.get("tips", []))

    return f"""You are a friendly, knowledgeable personal fitness coach having a conversation with your client.

Here is everything you know about them:

NAME       : {profile['name']}
AGE        : {profile['age']} years
HEIGHT     : {profile['height_cm']} cm
WEIGHT     : {profile['weight_kg']} kg
BMI        : {analysis['bmi']} ({analysis['category']})
EXPERIENCE : {profile['experience']}
STYLE      : {profile['style']}
GOAL       : {profile['goal']}
DAYS/WEEK  : {profile['days_per_week']}
EQUIPMENT  : {profile['equipment']}
COACH NOTE : {analysis['note']}

THEIR WORKOUT PLAN:
{plan_text}

THEIR PERSONALISED TIPS:
{tips_text}

Your job:
- Answer their fitness questions with full awareness of their profile and plan above
- Give specific, practical advice — not generic answers
- If they ask about modifying their plan, refer to the actual exercises in their plan
- Keep answers concise and motivating
- If they ask something unrelated to fitness, gently redirect them
- Never forget who they are — always personalise your answers to {profile['name']}
"""


# ─────────────────────────────────────────────
# STEP 8: The chat loop
# ─────────────────────────────────────────────

def start_chat(profile, analysis, data):
    system_prompt = build_system_prompt(profile, analysis, data)

    # conversation_history holds the full chat so the model remembers everything
    conversation_history = []

    console.print("\n")
    console.print(Panel(
        f"[bold green]Chat started![/bold green]\n"
        f"Ask me anything about your workout plan, exercises, nutrition, recovery — anything fitness related.\n"
        f"[dim]Type [bold]'quit'[/bold] or [bold]'exit'[/bold] to end the chat.[/dim]",
        title="💬  Personal Fitness Coach",
        border_style="green"
    ))

    while True:
        # Get user input
        console.print("\n[bold cyan]You:[/bold cyan] ", end="")
        user_input = input().strip()

        if not user_input:
            continue

        if user_input.lower() in ("quit", "exit", "bye", "q"):
            console.print("\n[bold green]Coach:[/bold green] Great work today! Stay consistent and you'll see results. 💪\n")
            break

        # Add user message to history
        conversation_history.append({
            "role": "user",
            "content": user_input
        })

        # Call Ollama with full history + system context
        console.print("\n[bold green]Coach:[/bold green] ", end="")

        try:
            full_reply = ""

            # Stream the reply word-by-word for a natural feel
            stream = ollama.chat(
                model="llama3.2:3b",
                messages=[
                    {"role": "system", "content": system_prompt},
                    *conversation_history          # full history every time
                ],
                stream=True
            )

            for chunk in stream:
                text = chunk["message"]["content"]
                print(text, end="", flush=True)
                full_reply += text

            print()  # newline after response

            # Add assistant reply to history so next message has full context
            conversation_history.append({
                "role": "assistant",
                "content": full_reply
            })

        except Exception as e:
            console.print(f"\n[red]❌  Error: {e}[/red]")
            console.print("   Make sure Ollama is still running.")
            # Remove the failed message from history
            conversation_history.pop()


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    # 1. Get profile
    profile  = get_user_profile()
    analysis = analyse_profile(profile)

    console.print(f"\n[bold]📊  Profile analysis:[/bold]")
    console.print(f"    BMI  :  {analysis['bmi']}  →  {analysis['category']}")
    console.print(f"    Note :  {analysis['note']}")

    # 2. Generate plan
    prompt = build_plan_prompt(profile, analysis)
    data   = generate_plan(prompt)
    if not data:
        return

    # 3. Display tables
    display_terminal_tables(profile, analysis, data)

    # 4. Save to Excel
    console.print("\n")
    if input("Save to Excel file? (y/n): ").strip().lower() == "y":
        filename = save_excel(profile, analysis, data)
        console.print(f"\n[bold green]✅  Saved:[/bold green] {filename}\n")

    # 5. Start chat
    console.print("\n")
    if input("Start the fitness chat assistant? (y/n): ").strip().lower() == "y":
        start_chat(profile, analysis, data)


if __name__ == "__main__":
    main()
